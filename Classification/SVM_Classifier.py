# -*- coding: utf-8 -*-
"""
Created on Sat Aug 24 11:45:19 2019

@author: Salim
"""

import pandas as pd
import numpy as np
from matplotlib import pyplot as plt
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import MinMaxScaler
from sklearn import svm
from sklearn.model_selection import GridSearchCV
import warnings
from sklearn.metrics import roc_curve
from sklearn import metrics
import openpyxl
from openpyxl import Workbook
from datetime import datetime
from sklearn.externals import joblib
import os

warnings.filterwarnings("ignore")

'''
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn import datasets
from sklearn.datasets import load_breast_cancer

def gen_dataset(n_samples=100000,n_features=30,n_classes=2,random_state=123):
    X,y=datasets.make_classification(
            n_features=n_features,
            n_samples=n_samples,
            n_informative=int(0.6*n_features),
            n_redundant=int(0.1*n_features),
            n_classes=n_classes,
            random_state=random_state)
    return(X,y)
    
X,y=gen_dataset(n_samples=100000,n_features=30,n_classes=2)

data=load_breast_cancer()
X=pd.DataFrame(X,columns=data.feature_names)
y=pd.DataFrame(y,columns=['y'])

dev_data,X_test,dev_data_y,y_test=train_test_split(X,y,stratify=y,test_size=0.5,random_state=42)
oot_data,X_test1,oot_data_y,y_test1=train_test_split(X_test,y_test,test_size=0.5,random_state=42)
pdv_data,psi_data,pdv_data_y,psi_data_y=train_test_split(X_test1,y_test1,test_size=0.5,random_state=42)

dev_data=pd.concat([pd.DataFrame(dev_data),pd.DataFrame(dev_data_y)],axis=1)

oot_data=pd.concat([pd.DataFrame(oot_data),pd.DataFrame(oot_data_y)],axis=1)

pdv_data=pd.concat([pd.DataFrame(pdv_data),pd.DataFrame(pdv_data_y)],axis=1)

psi_data=pd.concat([pd.DataFrame(psi_data),pd.DataFrame(psi_data_y)],axis=1)

cols_list=[col for col in dev_data.columns if col not in dependent]

'''

class SVM_Classifier():
    """
    An automated Support Vector Machine Classifier.

Parameters
----------
dev_data : dataframe, optional (default=None)
    Specify the name of your Development dataset 
    (as assigned to the dataframe of the .csv file).

oos_data : dataframe, optional (default=None)
    Specify the name of your Out of Sample dataset 
    (as assigned to the dataframe of the .csv file).
    
oot_data : dataframe, optional (default=None)
    Specify the name of your Out of Time dataset 
    (as assigned to the dataframe of the .csv file).
    
pdv_data : dataframe, optional (default=None)
    Specify the name of your Pre Deployment dataset 
    (as assigned to the dataframe of the .csv file).
    
psi_data : dataframe, optional (default=None)
    Specify the name of your Population Stability Index dataset 
    (as assigned to the dataframe of the .csv file).
    The PSI data would not have the target variable column,
    as the values for it are to be predicted.

dependent : string, optional (default="None")
    The name of the target variable whose values are to be predicted. 
    True values for this variable are present in all datasets except the PSI data.
    Values for this variable are predicted and compared to create necessary 
    metrics and also to analyse the performance of the PSI data.
    
norm_flag : bool, optional (default=False)
    Whether you want to normalize the continuous variables of your data.
    If True, it uses MinMaxScaler() to normalize all the continuous variables
    and brings them in the range of 0-1. If False, it develops the 
    model on the original values of the data. 

split_flag : bool, optional (default=False)
    Whether you want to split your Development data into Development and 
    OOS data. If True it splits the data into development and OOS
    using stratified sampling. If False, it uses the entire development 
    data without splitting.
    
split_fraction : float, optional (default=0.0)
    The fraction of split of the OOS data. It's value is passed
    only if the split_flag is set to be True.

hyperopt_flag : bool, optional (default=False)
    Whether you want to tune your hyperparameters. If True,
    it uses Grid Search CV to find
    the best parameters. If False it uses the default parameters. 

location : string, optional (default="None")
    The path of the folder where you want your results to be saved.
    
project_name : string, optional (default="None")
    The name of the project by which you want your results to be saved.

cols_list : list, optional (default="None")
    List of names of columns for which you want the Support Vector Machine Classifier
    module to be built upon. This column list also contains
    the name of the dependent variable.



Methods
----------
__init__ : Initializes all the variables when an 
    object of the class is called. It has all the above 
    mentioned parameters along with their default values.

clear_columns : Creates a dataframe for only those columns 
    which are specified in the parameter 'cols_list'. Rest other variables
    are not considered while executing the code. It does this for all the datasets.
    
split_train_test : Splits the 'dev_data' into 'train_data' and 'oos_data',
    if the 'split_flag' is True. It determines the size of the oos_data by the
    value of 'split_fraction'. The splitting is done by using stratified sampling.

split_X_Y : Splits the dataframe into X and Y. Y is the dataframe containing the
    target variable and X is the dataframe containing all other variables. This function
    is used by the function 'split_train_test', before it splits the data.

combine_X_Y : Combines the previously created dataframes (X and Y) into a single
    dataframe, after splitting has been performed.

process_data : Combines all the dataframes and does some preprocessing of the categorical
    variables. It converts all the categorical columns into uppercase and strips all the blank
    spaces, if there are any. It then encodes all categorical variables by using
    function 'convert_categorical'. After the conversion is done, the datasets are separated 
    again. Combining of datasets is done so that if the test datasets have more/less 
    categories as compared to the train data, the code will give an error. 

convert_categorical : Creates dummies for all categorical variables for the entered
    dataframe. It creates n-1 dummy columns for a column which has n categories and
    returns the new dataframe.

normalize : Normalizes all the continuous variables for all the dataframes.
    If True, it uses MinMaxScaler() to normalize all the continuous variables
    and brings them in the range of 0-1. If False, it keeps the dataframes as it is.

SVM_Train : Fits the Support Vector Machine Classifier model on the training data
    and returns the trained model.

hyperopt : Tunes the hyperparameters and chooses the best parameters from the default
    pool of values. If True it uses Grid Search CV to find the 
    best parameters and returns the model fitted with the best parameters. If False 
    it uses the default parameters and returns the original model.
    
create_ks_score : Uses the model to make predictions on the training data and also 
    calculates their probabilities. It sorts the probabilities in descending order
    and divides the data into 10 deciles and returns the KS table for the train data.
    
create_ks_table : Scores the probabilities of the specified dataset, sorts the
    probabilities in descending order and divides the data into 10 deciles using 
    the minimum and maximum probabilities of the training data bands as lower 
    and upper limit and returns the KS table for the specified data.

psi_calculation : Scores the probabilities for the 'psi_data', sorts the
    probabilities in descending order and divides the data into 10 deciles using 
    the minimum and maximum probabilities of the training data bands as lower 
    and upper limit. It performs certain calculations to finally create the psi table 
    and returns it.

save_files : Calls all other functions. It calculates the univariate distribution
    for each dataset and the correlation coefficient among all indepedent variables. 
    It saves all of the above mentioned results in an excel file along with the order 
    of variables which would be used to score new data in future (ORDER SHOULD NOT BE CHANGED).
    These things are saved at the location provided by the user with the
    project name (also provided by the user), without any human intervention. Along 
    with this, it also saves the KS tables for all datasets in the same excel file.
    It also saves the AUC charts created for all time periods (in the excel file 
    and as a separate .png file) and the trained model in serialized form for future use
    in the same location without any human intervention. All the files in the provided 
    location are saved such that the file names are followed by the time of the system 
    at which the program is executed. This is done to avoid overwriting of the files 
    when the program is executed multiple times.
    
    """
    def __init__(self,dev_data=None,oos_data=None,oot_data=None,pdv_data=None,psi_data=None,dependent=None,norm_flag=False,split_flag=False,split_fraction=0,cols_list=None,hyperopt_flag=False,location=None,project_name=None):
        self.dev_data=dev_data
        self.oos_data=oos_data
        self.oot_data=oot_data
        self.pdv_data=pdv_data
        self.psi_data=psi_data
        self.dependent=dependent
        self.norm_flag=norm_flag
        self.split_flag=split_flag
        self.split_fraction=split_fraction
        self.cols_list=cols_list
        self.hyperopt_flag=hyperopt_flag
        self.location=location
        self.project_name=project_name
    
    def clear_columns(self):
        self.dev_data=self.dev_data[self.cols_list]
        if self.oos_data is not None:
            self.oos_data=self.oos_data[self.cols_list]
        self.oot_data=self.oot_data[self.cols_list]
        self.pdv_data=self.pdv_data[self.cols_list]
        self.cols_list.remove(self.dependent)
        self.psi_data=self.psi_data[self.cols_list]
         
    def split_train_test(self):
        self.clear_columns()
        if self.split_flag==True:
            data1,data2=self.split_X_Y(self.dev_data)
            df2X,df3X,df2Y,df3Y = train_test_split(data1,data2,stratify=data2, test_size=self.split_fraction, random_state=42)
            self.train_data=self.combine_X_Y(df2X,df2Y)
            self.train_data.reset_index(inplace=True,drop=True)
            self.oos_data=self.combine_X_Y(df3X,df3Y)
            self.oos_data.reset_index(inplace=True,drop=True)
            self.copy_train=self.train_data
            self.copy_oos=self.oos_data
            self.copy_oot=self.oot_data
            self.copy_pdv=self.pdv_data
            self.copy_psi=self.psi_data
            
        elif self.split_flag==False:
            self.train_data=self.dev_data
            self.oos_data=self.oos_data
            self.copy_train=self.train_data
            self.copy_oos=self.oos_data
            self.copy_oot=self.oot_data
            self.copy_pdv=self.pdv_data
            self.copy_psi=self.psi_data
        
    def split_X_Y(self,df):
        X=df.drop([self.dependent],axis=1)
        Y=pd.DataFrame(df,columns=[self.dependent])
        return X,Y

    def combine_X_Y(self,df1,df2):
        df3=pd.concat([df1,pd.DataFrame(df2,columns=[self.dependent])],axis=1)
        return df3
    
    def process_data(self):
        if self.oos_data is not None:
            
            self.Y_train=pd.DataFrame(self.train_data[self.dependent],columns=[self.dependent])
            self.oos_Y=pd.DataFrame(self.oos_data[self.dependent],columns=[self.dependent])
            self.oot_Y=pd.DataFrame(self.oot_data[self.dependent],columns=[self.dependent])
            self.pdv_Y=pd.DataFrame(self.pdv_data[self.dependent],columns=[self.dependent])
            
            self.train_data['dummy']=0
            self.oos_data['dummy']=1
            self.oot_data['dummy']=2
            self.pdv_data['dummy']=3
            self.psi_data['dummy']=4
            
            combine=pd.concat([self.train_data[self.cols_list+['dummy']],self.oos_data[self.cols_list+['dummy']],self.oot_data[self.cols_list+['dummy']],self.pdv_data[self.cols_list+['dummy']],self.psi_data[self.cols_list+['dummy']]])
           
            combine[combine.select_dtypes(include=np.number).columns] = combine[combine.select_dtypes(include=np.number).columns].fillna(value=99999999)
            combine[combine.select_dtypes(exclude=np.number).columns] = combine[combine.select_dtypes(exclude=np.number).columns].fillna(value="MISSING")
            combine.loc[:,combine.dtypes=='O']=combine.loc[:,combine.dtypes=='O'].apply(lambda x: x.astype(str).str.upper()) 
            combine.loc[:,combine.dtypes=='O']=combine.loc[:,combine.dtypes=='O'].apply(lambda x: x.astype(str).str.strip()) 
            
            ohe=self.convert_categorical(combine,self.dependent)
            
            self.train_data=ohe[ohe['dummy']==0]
            self.oos_data=ohe[ohe['dummy']==1]
            self.oot_data=ohe[ohe['dummy']==2]
            self.pdv_data=ohe[ohe['dummy']==3]
            self.psi_data=ohe[ohe['dummy']==4]
            
            self.train_data.drop(['dummy'],axis=1,inplace=True)
            self.oos_data.drop(['dummy'],axis=1,inplace=True)
            self.oot_data.drop(['dummy'],axis=1,inplace=True)
            self.pdv_data.drop(['dummy'],axis=1,inplace=True)
            self.psi_data.drop(['dummy'],axis=1,inplace=True)
            
        else:
            
            self.Y_train=pd.DataFrame(self.train_data[self.dependent],columns=[self.dependent])
            self.oot_Y=pd.DataFrame(self.oot_data[self.dependent],columns=[self.dependent])
            self.pdv_Y=pd.DataFrame(self.pdv_data[self.dependent],columns=[self.dependent])
            
            self.train_data['dummy']=0
            self.oot_data['dummy']=1
            self.pdv_data['dummy']=2
            self.psi_data['dummy']=3
            
            combine=pd.concat([self.train_data[self.cols_list+['dummy']],self.oot_data[self.cols_list+['dummy']],self.pdv_data[self.cols_list+['dummy']],self.psi_data[self.cols_list+['dummy']]])
            
            combine[combine.select_dtypes(include=np.number).columns] = combine[combine.select_dtypes(include=np.number).columns].fillna(value=99999999)
            combine[combine.select_dtypes(exclude=np.number).columns] = combine[combine.select_dtypes(exclude=np.number).columns].fillna(value="MISSING")
            combine.loc[:,combine.dtypes=='O']=combine.loc[:,combine.dtypes=='O'].apply(lambda x: x.astype(str).str.upper()) 
            combine.loc[:,combine.dtypes=='O']=combine.loc[:,combine.dtypes=='O'].apply(lambda x: x.astype(str).str.strip()) 
            
            ohe=self.convert_categorical(combine,self.dependent)
        
            self.train_data=ohe[ohe['dummy']==0]
            self.oot_data=ohe[ohe['dummy']==1]
            self.pdv_data=ohe[ohe['dummy']==2]
            self.psi_data=ohe[ohe['dummy']==3]
    
            self.train_data.drop(['dummy'],axis=1,inplace=True)
            self.oot_data.drop(['dummy'],axis=1,inplace=True)
            self.pdv_data.drop(['dummy'],axis=1,inplace=True)
            self.psi_data.drop(['dummy'],axis=1,inplace=True)
            
    def convert_categorical(self, df, target_varname):
        ohe_df=pd.get_dummies(df,columns=df.select_dtypes(exclude=np.number).columns,drop_first=True)
        return ohe_df
    
    def normalize(self):
        if self.norm_flag==True:
            minmax = MinMaxScaler()
            self.train_data = pd.DataFrame(minmax.fit_transform(self.train_data.select_dtypes(include=np.number)),columns=self.train_data.columns)
            if self.oos_data is not None:    
                self.oos_data = pd.DataFrame(minmax.transform(self.oos_data.select_dtypes(include=np.number)),columns=self.oos_data.columns)
            self.oot_data = pd.DataFrame(minmax.transform(self.oot_data.select_dtypes(include=np.number)),columns=self.oot_data.columns)
            self.pdv_data = pd.DataFrame(minmax.transform(self.pdv_data.select_dtypes(include=np.number)),columns=self.pdv_data.columns)
            self.psi_data = pd.DataFrame(minmax.transform(self.psi_data.select_dtypes(include=np.number)),columns=self.psi_data.columns)
            
        elif self.norm_flag==False:
            self.train_data=self.train_data.select_dtypes(include=np.number)
            if self.oos_data is not None:
                self.oos_data=self.oos_data.select_dtypes(include=np.number)
            self.oot_data=self.oot_data.select_dtypes(include=np.number)
            self.pdv_data=self.pdv_data.select_dtypes(include=np.number)
            self.psi_data=self.psi_data.select_dtypes(include=np.number)

    def SVM_Train(self):
        self.clf=svm.SVC(C=1.0,gamma='scale', kernel='rbf', probability=True)
        self.clf.fit(self.train_data,self.Y_train)
        return self.clf
    
    def hyperopt(self):
        if self.hyperopt_flag==True:
            param_grid = {'C': [0.1,0.5,1,3,5],'kernel': ['linear','rbf'], 'gamma':[0.01,0.05,0.1,0.5,1.0]}
            self.clf_cv=GridSearchCV(self.clf, param_grid)
            self.clf_cv.fit(self.train_data, self.Y_train)
            print(self.clf_cv.best_params_)
            self.best_grid = self.clf_cv.best_estimator_
            return self.best_grid
        elif self.hyperopt_flag==False: 
            return self.clf
    
    def create_ks_score(self):
        prob=self.model.predict_proba(self.train_data)[:,1]
        score=pd.DataFrame(prob,columns=['SCORE']) 
        score['DECILE']=pd.qcut(score['SCORE'].rank(method='first'),10,labels=range(10,0,-1))    
        score['DECILE']=score['DECILE'].astype(float)
        self.Y_train.index=score.index  
        score['TARGET']=pd.DataFrame(self.Y_train)
        score['NON TARGET']=1-score['TARGET']
        score.sort_values(by=['SCORE'],ascending=False,inplace=True)
        score_min=score.groupby(['DECILE'],as_index=False).min()
        score_max=score.groupby(['DECILE'],as_index=False).max()
        score=score.groupby(['DECILE'],as_index=False).sum()
        score['TOTAL']=score['TARGET']+score['NON TARGET']
        score_min.index=score.index
        score_max.index=score.index
        score['Minimum Probability']=score_min['SCORE']
        score['Maximum Probability']=score_max['SCORE']
        score['BANDS']=">= "+np.round(score['Minimum Probability'],3).astype("str") +" AND < "+np.round(score['Maximum Probability'],3).astype("str")
        score.loc[0,"BANDS"]=">= "+np.round(score.loc[0,"Minimum Probability"],3).astype("str")
        score.loc[9,"BANDS"]="< "+np.round(score.loc[9,"Maximum Probability"],3).astype("str")
        score['Cumulative Target']=score['TARGET'].cumsum()
        score['Cumulative Non Target']=score['NON TARGET'].cumsum()
        score['Cumulative Total']=score['TOTAL'].cumsum()
        score['Population%']=np.round(score['TOTAL']/score['TOTAL'].sum()*100,2)
        score['BadCaptured%']=np.round(score['TARGET']/score['TARGET'].sum()*100,2)
        score['CumulativeTarget%']=np.round(score['Cumulative Target']/score['TARGET'].sum()*100,2)
        score['CumulativeNonTarget%']=np.round(score['Cumulative Non Target']/score['NON TARGET'].sum()*100,2)
        score['KS']=np.abs(np.round(score['CumulativeTarget%']-score['CumulativeNonTarget%'],2))
        score.index=score['BANDS']
        del score['SCORE'],score['DECILE'],score['BANDS']
        return score
        
    def create_ks_table(self,X_test,truth,dev_ks_table):
        prob=self.model.predict_proba(X_test)[:,1]
        score=pd.DataFrame(prob,columns=['SCORE'])
        min_prob=list(dev_ks_table['Minimum Probability'])
        score['Bands']=np.where(score['SCORE']>=min_prob[0],dev_ks_table.index[0],0)
        score['Bands']=np.where(((score['SCORE']>=min_prob[1])&(score['SCORE']<min_prob[0])),dev_ks_table.index[1],score['Bands'])
        score['Bands']=np.where(((score['SCORE']>=min_prob[2])&(score['SCORE']<min_prob[1])),dev_ks_table.index[2],score['Bands'])
        score['Bands']=np.where(((score['SCORE']>=min_prob[3])&(score['SCORE']<min_prob[2])),dev_ks_table.index[3],score['Bands'])
        score['Bands']=np.where(((score['SCORE']>=min_prob[4])&(score['SCORE']<min_prob[3])),dev_ks_table.index[4],score['Bands'])
        score['Bands']=np.where(((score['SCORE']>=min_prob[5])&(score['SCORE']<min_prob[4])),dev_ks_table.index[5],score['Bands'])
        score['Bands']=np.where(((score['SCORE']>=min_prob[6])&(score['SCORE']<min_prob[5])),dev_ks_table.index[6],score['Bands'])
        score['Bands']=np.where(((score['SCORE']>=min_prob[7])&(score['SCORE']<min_prob[6])),dev_ks_table.index[7],score['Bands'])
        score['Bands']=np.where(((score['SCORE']>=min_prob[8])&(score['SCORE']<min_prob[7])),dev_ks_table.index[8],score['Bands'])
        score['Bands']=np.where((score['SCORE']<min_prob[8]),dev_ks_table.index[9],score['Bands']) 
        truth.index=score.index
        score['TARGET']=pd.DataFrame(truth)
        score['NON TARGET']=1-score['TARGET']
        score=score.groupby(['Bands'],as_index=False).sum()
        score.sort_index(ascending=False,inplace=True)
        score['TOTAL']=score['TARGET']+score['NON TARGET']
        score['Cumulative Target']=score['TARGET'].cumsum()
        score['Cumulative Non Target']=score['NON TARGET'].cumsum()
        score['Cumulative Total']=score['TOTAL'].cumsum()
        score['Population%']=np.round(score['TOTAL']/score['TOTAL'].sum()*100,2)
        score['BadCaptured%']=np.round(score['TARGET']/score['TARGET'].sum()*100,2)
        score['CumulativeTarget%']=np.round(score['Cumulative Target']/score['TARGET'].sum()*100,2)
        score['CumulativeNonTarget%']=np.round(score['Cumulative Non Target']/score['NON TARGET'].sum()*100,2)
        score['KS']=np.abs(np.round(score['CumulativeTarget%']-score['CumulativeNonTarget%'],2))
        score.index=score['Bands']
        del score['SCORE'],score['Bands']
        return score
        
    def psi_calculation(self,X_test,ks):
        prob=self.model.predict_proba(X_test)[:,1]
        score=pd.DataFrame(prob,columns=['SCORE'])
        min_prob=list(ks['Minimum Probability'])
        score['Bands']=np.where(score['SCORE']>=min_prob[0],ks.index[0],0)
        score['Bands']=np.where(((score['SCORE']>=min_prob[1])&(score['SCORE']<min_prob[0])),ks.index[1],score['Bands'])
        score['Bands']=np.where(((score['SCORE']>=min_prob[2])&(score['SCORE']<min_prob[1])),ks.index[2],score['Bands'])
        score['Bands']=np.where(((score['SCORE']>=min_prob[3])&(score['SCORE']<min_prob[2])),ks.index[3],score['Bands'])
        score['Bands']=np.where(((score['SCORE']>=min_prob[4])&(score['SCORE']<min_prob[3])),ks.index[4],score['Bands'])
        score['Bands']=np.where(((score['SCORE']>=min_prob[5])&(score['SCORE']<min_prob[4])),ks.index[5],score['Bands'])
        score['Bands']=np.where(((score['SCORE']>=min_prob[6])&(score['SCORE']<min_prob[5])),ks.index[6],score['Bands'])
        score['Bands']=np.where(((score['SCORE']>=min_prob[7])&(score['SCORE']<min_prob[6])),ks.index[7],score['Bands'])
        score['Bands']=np.where(((score['SCORE']>=min_prob[8])&(score['SCORE']<min_prob[7])),ks.index[8],score['Bands'])
        score['Bands']=np.where((score['SCORE']<min_prob[8]),ks.index[9],score['Bands'])
        score=score.groupby(['Bands'],as_index=False).count()
        score.sort_values(by=['Bands'],ascending=False,inplace=True)
        score.index=ks.index
        score['Expected%']=np.round(score['SCORE']/score['SCORE'].sum()*100,2)
        score['Actual%']=ks['Population%']
        score['Ac-Ex']=score['Actual%']-score['Expected%']
        score['ln(Ac/Ex)']=np.log(score['Actual%']/score['Expected%'])
        score['index']=score['Ac-Ex']*score['ln(Ac/Ex)']
        del score['SCORE'],score['Bands']
        return score
        
    def save_files(self):
        
        self.split_train_test()
        self.process_data()
        self.normalize()
        self.clf=self.SVM_Train()
        self.model=self.hyperopt()
        self.dev_ks_table=self.create_ks_score()
        if self.oos_data is not None:
            self.ks_table_oos=self.create_ks_table(self.oos_data,self.oos_Y,self.dev_ks_table)
        else:
            self.ks_table_oos=None
        self.ks_table_oot=self.create_ks_table(self.oot_data,self.oot_Y,self.dev_ks_table)
        self.ks_table_pdv=self.create_ks_table(self.pdv_data,self.pdv_Y,self.dev_ks_table)
        self.psi_table=self.psi_calculation(self.psi_data,self.dev_ks_table)
        
        time=datetime.now()
        book=Workbook()  
        path=os.path.join(self.location,"Performance_"+str(self.project_name)+"_"+time.strftime("%d-%m-%Y %H-%M-%S")+".xlsx")
        writer = pd.ExcelWriter(path, engine='openpyxl')
        writer.book=book
        self.dev_ks_table.to_excel(writer,sheet_name="Development Performance")
        pd.concat([self.copy_train.describe(percentiles=[0.005,0.01,0.025,0.05,0.1,0.2,0.25,0.3,0.4,0.5,0.6,0.7,0.75,0.8,0.9,0.975,0.99]).T,pd.DataFrame(self.copy_train.isnull().sum(),columns=['Missing Values'])],axis=1).to_excel(writer,sheet_name="Development Variables Distribution")
        if self.ks_table_oos is not None:
            self.ks_table_oos.to_excel(writer,sheet_name="OOS Performance")
            pd.concat([self.copy_oos.describe(percentiles=[0.005,0.01,0.025,0.05,0.1,0.2,0.25,0.3,0.4,0.5,0.6,0.7,0.75,0.8,0.9,0.975,0.99]).T,pd.DataFrame(self.copy_oos.isnull().sum(),columns=['Missing Values'])],axis=1).to_excel(writer,sheet_name="OOS Variables Distribution")
        self.ks_table_oot.to_excel(writer,sheet_name="OOT Performance")
        pd.concat([self.copy_oot.describe(percentiles=[0.005,0.01,0.025,0.05,0.1,0.2,0.25,0.3,0.4,0.5,0.6,0.7,0.75,0.8,0.9,0.975,0.99]).T,pd.DataFrame(self.copy_oot.isnull().sum(),columns=['Missing Values'])],axis=1).to_excel(writer,sheet_name="OOT Variables Distribution")
        self.ks_table_pdv.to_excel(writer,sheet_name="PDV Performance")
        pd.concat([self.copy_pdv.describe(percentiles=[0.005,0.01,0.025,0.05,0.1,0.2,0.25,0.3,0.4,0.5,0.6,0.7,0.75,0.8,0.9,0.975,0.99]).T,pd.DataFrame(self.copy_pdv.isnull().sum(),columns=['Missing Values'])],axis=1).to_excel(writer,sheet_name="PDV Variables Distribution")
        self.psi_table.to_excel(writer, sheet_name="PSI Calculation")
        pd.concat([self.copy_psi.describe(percentiles=[0.005,0.01,0.025,0.05,0.1,0.2,0.25,0.3,0.4,0.5,0.6,0.7,0.75,0.8,0.9,0.975,0.99]).T,pd.DataFrame(self.copy_psi.isnull().sum(),columns=['Missing Values'])],axis=1).to_excel(writer,sheet_name="PSI Variables Distribution")
        self.train_data.corr().to_excel(writer,sheet_name="Correlation Matrix")
        auc=metrics.roc_auc_score(self.Y_train,self.model.predict_proba(self.train_data)[::,1])
        fpr, tpr, _ = roc_curve(self.Y_train, self.model.predict_proba(self.train_data)[::,1])
        plt.plot([0, 1], [0, 1],'r--')
        plt.grid()
        plt.title("ROC Curve", fontsize=14)
        plt.ylabel('True Positive Rate',fontsize=12)
        plt.xlabel('False Positive Rate',fontsize=12)
        plt.plot(fpr,tpr,marker='.')
        if self.oos_data is not None:
            auc1=metrics.roc_auc_score(self.oos_Y,self.model.predict_proba(self.oos_data)[::,1])
            fpr, tpr, _ = roc_curve(self.oos_Y, self.model.predict_proba(self.oos_data)[::,1])
            plt.plot(fpr,tpr,marker='.')
            auc2=metrics.roc_auc_score(self.oot_Y,self.model.predict_proba(self.oot_data)[::,1])
            fpr, tpr, _ = roc_curve(self.oot_Y, self.model.predict_proba(self.oot_data)[::,1])
            plt.plot(fpr,tpr,marker='.') 
            auc3=metrics.roc_auc_score(self.pdv_Y,self.model.predict_proba(self.pdv_data)[::,1])
            fpr, tpr, _ = roc_curve(self.pdv_Y, self.model.predict_proba(self.pdv_data)[::,1])
            plt.plot(fpr,tpr,marker='.') 
            plt.legend(["Base AUC=0.5","Train AUC=%.3f"%auc,"OOS AUC=%.3f"%auc1,"OOT AUC=%.3f"%auc2,"PDV AUC=%.3f"%auc3])
        else:
            auc2=metrics.roc_auc_score(self.oot_Y,self.model.predict_proba(self.oot_data)[::,1])
            fpr, tpr, _ = roc_curve(self.oot_Y, self.model.predict_proba(self.oot_data)[::,1])
            plt.plot(fpr,tpr,marker='.') 
            auc3=metrics.roc_auc_score(self.pdv_Y,self.model.predict_proba(self.pdv_data)[::,1])
            fpr, tpr, _ = roc_curve(self.pdv_Y, self.model.predict_proba(self.pdv_data)[::,1])
            plt.plot(fpr,tpr,marker='.') 
            plt.legend(["Base AUC=0.5","Train AUC=%.3f"%auc,"OOT AUC=%.3f"%auc2,"PDV AUC=%.3f"%auc3])
        book.remove(book['Sheet'])
        writer.save()
        writer.close()
        wb=openpyxl.load_workbook(path)
        ws=wb.create_sheet("AUC Chart")
        plt.savefig(os.path.join(self.location,"auc_chart"+time.strftime("%d-%m-%Y %H-%M-%S")+".png"),dpi=150)
        img=openpyxl.drawing.image.Image(os.path.join(self.location,"auc_chart"+time.strftime("%d-%m-%Y %H-%M-%S")+".png"))
        ws.add_image(img)
        wb.save(path)
        wb.close()
        joblib.dump(self.model,os.path.join(self.location,self.project_name+"-"+time.strftime("%d-%m-%Y %H-%M-%S")+".joblib.dat"))
        del self.copy_train,self.copy_oos,self.copy_oot,self.copy_pdv,self.copy_psi
        return self.train_data,self.oos_data,self.oot_data,self.pdv_data,self.psi_data,self.dev_ks_table,self.ks_table_oos,self.ks_table_oot,self.ks_table_pdv,self.psi_table
