# -*- coding: utf-8 -*-
"""
Created on Mon Aug 12 13:13:34 2019

@author: Salim
"""

import pandas as pd
import numpy as np
from matplotlib import pyplot as plt
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import GridSearchCV
from sklearn.metrics import r2_score
from sklearn.metrics import mean_squared_error
from sklearn.tree import export_graphviz
import warnings
from subprocess import call
import openpyxl
from openpyxl import Workbook
import operator
from datetime import datetime
from sklearn.externals import joblib
import os

warnings.filterwarnings("ignore")

class Random_Forest_Regressor():
    """
An automated Random Forest Regressor.
Parameters
----------
dev_data : dataframe, optional (default=None)
    Specify the name of your Development dataset 
    (as assigned to the dataframe of the .csv file).

oos_data : dataframe, optional (default=None)
    Specify the name of your Out of Sample dataset 
    (as assigned to the dataframe of the .csv file).
    
oot_data : dataframe, optional (default=None)
    Specify the name of your Out of Time dataset 
    (as assigned to the dataframe of the .csv file).
    
pdv_data : dataframe, optional (default=None)
    Specify the name of your Pre Deployment dataset 
    (as assigned to the dataframe of the .csv file).
    
psi_data : dataframe, optional (default=None)
    Specify the name of your Population Stability Index dataset 
    (as assigned to the dataframe of the .csv file).
    The PSI data would not have the target variable column,
    as the values for it are to be predicted.

dependent : string, optional (default="None")
    The name of the target variable whose values are to be predicted. 
    True values for this variable are present in all datasets except the PSI data.
    Values for this variable are predicted and compared to create necessary 
    metrics and also to analyse the performance of the PSI data.

split_flag : bool, optional (default=False)
    Whether you want to split your Development data into Development and 
    OOS data. If True it splits the data into development and OOS
    using stratified sampling. If False, it uses the entire development 
    data without splitting.
    
split_fraction : float, optional (default=0.0)
    The fraction of split of the OOS data. It's value is passed
    only if the norm_flag is set to be True.

hyperopt_flag : bool, optional (default=False)
    Whether you want to tune your hyperparameters. If True,
    it uses Grid Search CV to find
    the best parameters. If False it uses the default parameters. 

location : string, optional (default="None")
    The path of the folder where you want your results to be saved.
    
project_name : string, optional (default="None")
    The name of the project by which you want your results to be saved.

cols_list : list, optional (default="None")
    List of names of columns for which you want the Random Forest
    module to be built upon. This column list also contains
    the name of the dependent variable.



Methods
----------
__init__ : Initializes all the variables when an 
    object of the class is called. It has all the above 
    mentioned parameters along with their default values.

clear_columns : Creates a dataframe for only those columns 
    which are specified in the parameter 'cols_list'. Rest other variables
    are not considered while executing the code. It does this for all the datasets.
    
split_train_test : Splits the 'dev_data' into 'train_data' and 'oos_data',
    if the 'split_flag' is True. It determines the size of the oos_data by the
    value of 'split_fraction'. The splitting is done by using stratified sampling.

split_X_Y : Splits the dataframe into X and Y. Y is the dataframe containing the
    target variable and X is the dataframe containing all other variables. This function
    is used by the function 'split_train_test', before it splits the data.

combine_X_Y : Combines the previously created dataframes (X and Y) into a single
    dataframe, after splitting has been performed.

process_data : Combines all the dataframes and does some preprocessing of the categorical
    variables. It converts all the categorical columns into uppercase and strips all the blank
    spaces, if there are any. It then encodes all categorical variables by using
    function 'convert_categorical'. After the conversion is done, the datasets are separated 
    again. Combining of datasets is done so that if the test datasets have more/less 
    categories as compared to the train data, the code will give an error. 

convert_categorical : Creates dummies for all categorical variables for the entered
    dataframe. It creates n-1 dummy columns for a column which has n categories and
    returns the new dataframe.

Random_Forest_Train : Fits the Random Forest Regressor model on the training data
    and returns the trained model.

hyperopt : Tunes the hyperparameters and chooses the best parameters from the default
    pool of values.If True it uses Grid Search CV to find the 
    best parameters and returns the model fitted with the best parameters. If False 
    it uses the default parameters and returns the original model.
    
reg_metrics : Scores the dataset and calculates R square, Adjusted R square, 
    Mean Squared Error and Root Mean Squared Error for it. It returns a table created
    out of the values of the above mentioned metrics. 

psi_calculation : Scores the training data using the model and creates bands of 20 
    percentiles each. It then scores the PSI data and gives the median of values in each
    bin, after classifying them as per the percentile bins created by the dependent 
    column of the training data. It returns a table created out of the median values of 
    the training and the psi data.

save_files : Calls all other functions. It calculates the univariate distribution
    for each dataset, feature importance for each variable (and sorts them in descending order),
    the correlation coefficient among all indepedent variables. It saves all of the 
    above mentioned results in an excel file along with the order of variables which 
    would be used to score new data in future (ORDER SHOULD NOT BE CHANGED).
    These things are saved at the location provided by the user with the
    project name (also provided by the user), without any human intervention. Along 
    with this, it also saves the Performance tables for all datasets in the same excel file.
    It also saves the PSI table and the PSI Performance Graph (in the excel file 
    and as a separate .png file) and the trained model in serialized form for future use
    (in the excel file and as a separate .png file) in the same location without any human intervention.
    All the files in the provided location are saved such that the file names 
    are followed by the time of the system at which the program is executed. This is 
    done to avoid overwriting of the files when the program is executed multiple times.                                                                

    """
    def __init__(self,dev_data=None,oos_data=None,oot_data=None,pdv_data=None,psi_data=None,dependent=None,split_flag=False,split_fraction=0,cols_list=None,hyperopt_flag=False,location=None,project_name=None):
        self.dev_data=dev_data
        self.oos_data=oos_data
        self.oot_data=oot_data
        self.pdv_data=pdv_data
        self.psi_data=psi_data
        self.dependent=dependent
        self.split_flag=split_flag
        self.split_fraction=split_fraction
        self.cols_list=cols_list
        self.hyperopt_flag=hyperopt_flag
        self.location=location
        self.project_name=project_name
    
    def clear_columns(self):
        self.dev_data=self.dev_data[self.cols_list]
        if self.oos_data is not None:
            self.oos_data=self.oos_data[self.cols_list]
        self.oot_data=self.oot_data[self.cols_list]
        self.pdv_data=self.pdv_data[self.cols_list]
        self.cols_list.remove(self.dependent)
        self.psi_data=self.psi_data[self.cols_list]
         
    def split_train_test(self):
        self.clear_columns()
        if self.split_flag==True:
            data1,data2=self.split_X_Y(self.dev_data)
            df2X,df3X,df2Y,df3Y = train_test_split(data1,data2,test_size=self.split_fraction, random_state=42)
            self.train_data=self.combine_X_Y(df2X,df2Y)
            self.train_data.reset_index(inplace=True)
            self.oos_data=self.combine_X_Y(df3X,df3Y)
            self.oos_data.reset_index(inplace=True)
            self.copy_train=self.train_data
            self.copy_oos=self.oos_data
            self.copy_oot=self.oot_data
            self.copy_pdv=self.pdv_data
            self.copy_psi=self.psi_data
            
        elif self.split_flag==False:
            self.train_data=self.dev_data
            self.oos_data=self.oos_data
            self.copy_train=self.train_data
            self.copy_oos=self.oos_data
            self.copy_oot=self.oot_data
            self.copy_pdv=self.pdv_data
            self.copy_psi=self.psi_data
        
    def split_X_Y(self,df):
        X=df.drop([self.dependent],axis=1)
        Y=pd.DataFrame(df,columns=[self.dependent])
        return X,Y

    def combine_X_Y(self,df1,df2):
        df3=pd.concat([df1,pd.DataFrame(df2,columns=[self.dependent])],axis=1)
        return df3
    
    def process_data(self):
        if self.oos_data is not None:
            
            self.train_data.loc[self.train_data[self.dependent]==0, self.dependent] = 0.001
            self.Y_train=pd.DataFrame(np.log(self.train_data[self.dependent]),columns=[self.dependent])
            self.oos_data.loc[self.oos_data[self.dependent]==0, self.dependent] = 0.001
            self.oos_Y=pd.DataFrame(np.log(self.oos_data[self.dependent]),columns=[self.dependent])
            self.oot_data.loc[self.oot_data[self.dependent]==0, self.dependent] = 0.001
            self.oot_Y=pd.DataFrame(np.log(self.oot_data[self.dependent]),columns=[self.dependent])
            self.pdv_data.loc[self.pdv_data[self.dependent]==0, self.dependent] = 0.001
            self.pdv_Y=pd.DataFrame(np.log(self.pdv_data[self.dependent]),columns=[self.dependent])
            
            self.train_data['dummy']=0
            self.oos_data['dummy']=1
            self.oot_data['dummy']=2
            self.pdv_data['dummy']=3
            self.psi_data['dummy']=4

            combine=pd.concat([self.train_data[self.cols_list+['dummy']],self.oos_data[self.cols_list+['dummy']],self.oot_data[self.cols_list+['dummy']],self.pdv_data[self.cols_list+['dummy']],self.psi_data[self.cols_list+['dummy']]])
            
#            combine[combine.select_dtypes(include=np.number).columns] = combine[combine.select_dtypes(include=np.number).columns].fillna(value=99999999)
            combine[combine.select_dtypes(exclude=np.number).columns] = combine[combine.select_dtypes(exclude=np.number).columns].fillna(value="MISSING")
            combine.loc[:,combine.dtypes=='O']=combine.loc[:,combine.dtypes=='O'].apply(lambda x: x.astype(str).str.upper()) 
            combine.loc[:,combine.dtypes=='O']=combine.loc[:,combine.dtypes=='O'].apply(lambda x: x.astype(str).str.strip()) 
          
            ohe=self.convert_categorical(combine)
            
            self.train_data=ohe[ohe['dummy']==0]
            self.oos_data=ohe[ohe['dummy']==1]
            self.oot_data=ohe[ohe['dummy']==2]
            self.pdv_data=ohe[ohe['dummy']==3]
            self.psi_data=ohe[ohe['dummy']==4]
            
            self.train_data.drop(['dummy'],axis=1,inplace=True)
            self.oos_data.drop(['dummy'],axis=1,inplace=True)
            self.oot_data.drop(['dummy'],axis=1,inplace=True)
            self.pdv_data.drop(['dummy'],axis=1,inplace=True)
            self.psi_data.drop(['dummy'],axis=1,inplace=True)
            
        else:
            
            self.train_data.loc[self.train_data[self.dependent]==0, self.dependent] = 0.001
            self.Y_train=pd.DataFrame(np.log(self.train_data[self.dependent]),columns=[self.dependent])
            self.oot_data.loc[self.oot_data[self.dependent]==0, self.dependent] = 0.001
            self.oot_Y=pd.DataFrame(np.log(self.oot_data[self.dependent]),columns=[self.dependent])
            self.pdv_data.loc[self.pdv_data[self.dependent]==0, self.dependent] = 0.001
            self.pdv_Y=pd.DataFrame(np.log(self.pdv_data[self.dependent]),columns=[self.dependent])
            
            self.train_data['dummy']=0
            self.oot_data['dummy']=1
            self.pdv_data['dummy']=2
            self.psi_data['dummy']=3
            
            combine=pd.concat([self.train_data[self.cols_list+['dummy']],self.oot_data[self.cols_list+['dummy']],self.pdv_data[self.cols_list+['dummy']],self.psi_data[self.cols_list+['dummy']]])
            
#            combine[combine.select_dtypes(include=np.number).columns] = combine[combine.select_dtypes(include=np.number).columns].fillna(value=99999999)
            combine[combine.select_dtypes(exclude=np.number).columns] = combine[combine.select_dtypes(exclude=np.number).columns].fillna(value="MISSING")
            combine.loc[:,combine.dtypes=='O']=combine.loc[:,combine.dtypes=='O'].apply(lambda x: x.astype(str).str.upper()) 
            combine.loc[:,combine.dtypes=='O']=combine.loc[:,combine.dtypes=='O'].apply(lambda x: x.astype(str).str.strip()) 
            
            ohe=self.convert_categorical(combine)
        
            self.train_data=ohe[ohe['dummy']==0]
            self.oot_data=ohe[ohe['dummy']==1]
            self.pdv_data=ohe[ohe['dummy']==2]
            self.psi_data=ohe[ohe['dummy']==3]
    
            self.train_data.drop(['dummy'],axis=1,inplace=True)
            self.oot_data.drop(['dummy'],axis=1,inplace=True)
            self.pdv_data.drop(['dummy'],axis=1,inplace=True)
            self.psi_data.drop(['dummy'],axis=1,inplace=True)
            
    def convert_categorical(self, df):
        ohe_df=pd.get_dummies(df,columns=df.select_dtypes(exclude=np.number).columns,drop_first=True)
        return ohe_df

    def Random_Forest_Train(self):
        self.forest=RandomForestRegressor(n_jobs=1)
        self.forest.fit(self.train_data,self.Y_train)
        return self.forest
    
    def hyperopt(self):
        if self.hyperopt_flag==True:
            param_grid = {'max_depth': [3,5,7,9],'n_estimators': [50,200,400,600]}
            self.forest_cv=GridSearchCV(self.forest, param_grid)
            self.forest_cv.fit(self.train_data, self.Y_train)
            print(self.forest_cv.best_params_)
            self.best_grid = self.forest_cv.best_estimator_
            return self.best_grid
        elif self.hyperopt_flag==False: 
            return self.forest
    
    def reg_metrics(self,X_test,Y_test):
        truth=np.exp(Y_test)
        pred=self.model.predict(X_test)
        Y_pred=np.exp(pred)
        metrics=['R squared','Adjusted R squared','Mean Square Error','RMSE']
        score=pd.DataFrame(metrics,columns=['Metrics'])
        r_squared=r2_score(truth,Y_pred)
        adjusted_r_squared=1 - (1-r_squared)*(len(Y_test)-1)/(len(Y_test)-X_test.shape[1]-1)
        mse=mean_squared_error(truth,Y_pred)
        rmse=np.sqrt(mse)
        values=[r_squared,adjusted_r_squared,mse,rmse]
        score['Values']=values
        return score
        
    def psi_calculation(self):
        score=np.exp(self.Y_train)
        score.columns=['Expected']
        values=[score.Expected.min(),score.Expected.quantile(0.2),score.Expected.quantile(0.4),score.Expected.quantile(0.6),score.Expected.quantile(0.8),score.Expected.max()]
        score['Bands'] = 1
        score.loc[((score['Expected']>=values[0])&(score['Expected']<values[1])), 'Bands'] = 1
        score.loc[((score['Expected']>=values[1])&(score['Expected']<values[2])), 'Bands'] = 2
        score.loc[((score['Expected']>=values[2])&(score['Expected']<values[3])), 'Bands'] = 3
        score.loc[((score['Expected']>=values[3])&(score['Expected']<values[4])), 'Bands'] = 4
        score.loc[((score['Expected']>=values[4])&(score['Expected']<values[5])), 'Bands'] = 5
        score=score.groupby(['Bands'],as_index=False).Expected.median()
        pred=self.model.predict(self.psi_data)
        actual=pd.DataFrame(np.exp(pred),columns=['Actual'])
        actual['Bands'] = 1
        actual.loc[((actual['Actual']>=values[0])&(actual['Actual']<values[1])), 'Bands'] = 1
        actual.loc[((actual['Actual']>=values[1])&(actual['Actual']<values[2])), 'Bands'] = 2
        actual.loc[((actual['Actual']>=values[2])&(actual['Actual']<values[3])), 'Bands'] = 3
        actual.loc[((actual['Actual']>=values[3])&(actual['Actual']<values[4])), 'Bands'] = 4
        actual.loc[((actual['Actual']>=values[4])&(actual['Actual']<values[5])), 'Bands'] = 5
        actual=actual.groupby(['Bands'],as_index=False).Actual.median()
        score['Actual']=actual['Actual']      
        score.loc[0,'Bands']=">= "+np.round(values[0],3).astype("str") +" AND < "+np.round(values[1],3).astype("str")
        score.loc[1,'Bands']=">= "+np.round(values[1],3).astype("str") +" AND < "+np.round(values[2],3).astype("str")
        score.loc[2,'Bands']=">= "+np.round(values[2],3).astype("str") +" AND < "+np.round(values[3],3).astype("str")
        score.loc[3,'Bands']=">= "+np.round(values[3],3).astype("str") +" AND < "+np.round(values[4],3).astype("str")
        score.loc[4,'Bands']=">= "+np.round(values[4],3).astype("str") +" AND < "+np.round(values[5],3).astype("str")
        return score
        
    
    def save_files(self):  
        self.split_train_test()
        self.process_data()
        self.forest=self.Random_Forest_Train()
        self.model=self.hyperopt()
        self.table_train=self.reg_metrics(self.train_data,self.Y_train)
        if self.oos_data is not None:
            self.table_oos=self.reg_metrics(self.oos_data,self.oos_Y)
        else:
            self.table_oos=None
        self.table_oot=self.reg_metrics(self.oot_data,self.oot_Y)
        self.table_pdv=self.reg_metrics(self.pdv_data,self.pdv_Y)
        self.table_psi=self.psi_calculation()
        time=datetime.now()
        book=Workbook()  
        path=os.path.join(self.location,"Performance_"+str(self.project_name)+"_"+time.strftime("%d-%m-%Y %H-%M-%S")+".xlsx")
        writer = pd.ExcelWriter(path, engine='openpyxl')
        writer.book=book
        self.table_train.to_excel(writer,sheet_name="Development Performance")
        pd.concat([self.copy_train.describe(percentiles=[0.005,0.01,0.025,0.05,0.1,0.2,0.25,0.3,0.4,0.5,0.6,0.7,0.75,0.8,0.9,0.975,0.99]).T,pd.DataFrame(self.copy_train.isnull().sum(),columns=['Missing Values'])],axis=1).to_excel(writer,sheet_name="Development Variables Distribution")
        if self.table_oos is not None:
            self.table_oos.to_excel(writer,sheet_name="OOS Performance")
            pd.concat([self.copy_oos.describe(percentiles=[0.005,0.01,0.025,0.05,0.1,0.2,0.25,0.3,0.4,0.5,0.6,0.7,0.75,0.8,0.9,0.975,0.99]).T,pd.DataFrame(self.copy_oos.isnull().sum(),columns=['Missing Values'])],axis=1).to_excel(writer,sheet_name="OOS Variables Distribution")
        self.table_oot.to_excel(writer,sheet_name="OOT Performance")
        pd.concat([self.copy_oot.describe(percentiles=[0.005,0.01,0.025,0.05,0.1,0.2,0.25,0.3,0.4,0.5,0.6,0.7,0.75,0.8,0.9,0.975,0.99]).T,pd.DataFrame(self.copy_oot.isnull().sum(),columns=['Missing Values'])],axis=1).to_excel(writer,sheet_name="OOT Variables Distribution")
        self.table_pdv.to_excel(writer,sheet_name="PDV Performance")
        pd.concat([self.copy_pdv.describe(percentiles=[0.005,0.01,0.025,0.05,0.1,0.2,0.25,0.3,0.4,0.5,0.6,0.7,0.75,0.8,0.9,0.975,0.99]).T,pd.DataFrame(self.copy_pdv.isnull().sum(),columns=['Missing Values'])],axis=1).to_excel(writer,sheet_name="PDV Variables Distribution")
        self.table_psi.to_excel(writer,sheet_name="PSI Performance")
        pd.concat([self.copy_psi.describe(percentiles=[0.005,0.01,0.025,0.05,0.1,0.2,0.25,0.3,0.4,0.5,0.6,0.7,0.75,0.8,0.9,0.975,0.99]).T,pd.DataFrame(self.copy_psi.isnull().sum(),columns=['Missing Values'])],axis=1).to_excel(writer,sheet_name="PSI Variables Distribution")
        fi=list(zip(self.train_data.columns,self.model.feature_importances_))
        fi.sort(key=operator.itemgetter(1),reverse=True)
        pd.DataFrame(fi,columns=['Variables','Feature Importance']).to_excel(writer,sheet_name="Feature Importances",index=False)
        self.train_data.corr().to_excel(writer,sheet_name="Correlation Matrix") 
        ax = plt.gca() 
        plt.title("PSI Performance", fontsize=14)
        plt.ylabel('Median Values',fontsize=12)
        plt.xlabel('Bands',fontsize=12)
        self.table_psi.plot(kind='line',x='Bands',y='Expected',ax=ax,color='green')
        self.table_psi.plot(kind='line',x='Bands',y='Actual', ax=ax,color='red')
        plt.setp(ax.get_xticklabels(), rotation=30, horizontalalignment='right')
        book.remove(book['Sheet'])
        writer.save()
        writer.close()
        wb=openpyxl.load_workbook(path)
        ws=wb.create_sheet("PSI Graph")
        plt.savefig(os.path.join(self.location,"psi_graph"+time.strftime("%d-%m-%Y %H-%M-%S")+".png"),dpi=150)
        img=openpyxl.drawing.image.Image(os.path.join(self.location,"psi_graph"+time.strftime("%d-%m-%Y %H-%M-%S")+".png"))
        ws.add_image(img)
        wb.save(path)
        wb.close()                
#        wb=openpyxl.load_workbook(path)
#        export_graphviz(self.model, out_file='tree.dot',rounded = True,feature_names = self.train_data.columns, proportion = False,precision = 2, filled = True)
#        call(['dot', '-Tpng', 'tree.dot', '-o', 'tree.png', '-Gdpi=600'])
#        ws=wb.create_sheet("Train Tree")
#        plt.figure(figsize = (14, 18))
#        plt.imshow(plt.imread('tree.png'))
#        plt.axis('off')
#        plt.savefig(os.path.join(self.location,"Train_Tree"+time.strftime("%d-%m-%Y %H-%M-%S")+".png"),dpi=600)
#        img=openpyxl.drawing.image.Image(os.path.join(self.location,"Train_Tree"+time.strftime("%d-%m-%Y %H-%M-%S")+".png"))
#        ws.add_image(img)
#        wb.save(path)
#        wb.close()
        joblib.dump(self.model,os.path.join(self.location,self.project_name+"-"+time.strftime("%d-%m-%Y %H-%M-%S")+".joblib.dat"))
        del self.copy_train,self.copy_oos,self.copy_oot,self.copy_pdv,self.copy_psi
        return self.train_data,self.oos_data,self.oot_data,self.pdv_data,self.psi_data,self.table_train,self.table_oos,self.table_oot,self.table_pdv,self.table_psi

dev_data=pd.read_csv('Regs_dev.csv')
oos_data=pd.read_csv('Regs_oot.csv')
oot_data=pd.read_csv('Regs_oot.csv')
pdv_data=pd.read_csv('Regs_pdv.csv')
psi_data=pd.read_csv('Regs_psi.csv')
        
cols_list=[' age', ' job ', ' marital ', ' education', ' default',
       ' housing', ' loan', ' contact', ' month', ' day_of_week', ' duration',
       ' campaign', ' pdays', ' previous', ' poutcome', ' emp_var_rate',
       ' cons_price_idx', ' cons_conf_idx', ' euribor3m', ' nr_employed']
project_name="Random_Forest_Regressor"
location="D:\Project 5\Random Forest Regressor"

model1=Random_Forest_Regressor(dev_data=dev_data,oos_data=oos_data,oot_data=oot_data,pdv_data=pdv_data,psi_data=psi_data,dependent=' duration',split_flag=False,hyperopt_flag=False,location=location,project_name=project_name,cols_list=cols_list)
train_final,oos_final,oot_final,pdv_final,psi_final,dev_ks_table,ks_table_oos,ks_table_oot,ks_table_pdv,psi_table=model1.save_files()